# -*- coding: utf-8 -*-
"""
Created on Tue Apr 21 16:50:38 2020

@author: ufooo
"""

import numpy as np
import pandas as pd


# 1. Kopf 교수님의 데이터에서 PCC6803의 whole genome sequence 얻기

pccgene = []
pccseq = ''
data = open('/Users/ufooo/OneDrive/문서/Python Scripts/PCC6803 gene info.txt','r')
line = data.readlines()
gspoint = line.index('ORIGIN\n')
for i in range(len(line)-gspoint-1):
    templist = line[i+gspoint].strip().split()[1:]
    for partseq in templist:
        pccseq = pccseq+partseq
data.close()

pccseq = pccseq.upper()

# 2. Kopf 교수님의 엑셀데이터에 접근하여 excel data를 training 할 수 있는 dataset으로 바꾸기

trlist = []
comp_dic = {'A':'T','G':'C','C':'G','T':'A'}
prolen = 50
df = pd.read_excel('/Users/ufooo/OneDrive/문서/Python Scripts/PCC6803 promoter reads info.xlsx',header = 1,sheet_name = 'S1 - All TUs')
numpydf = df.to_numpy()
for i in numpydf:
    if i[0] == 'Chr':
        if i[4] == '+':
            trlist.append((pccseq[i[2]-(prolen+1):i[2]-1],i[19]))
        else:
            compseq = ''
            seq = pccseq[i[2]:i[2]+prolen]
            for s in seq:
                compseq +=comp_dic[s]
            revcomp_seq = compseq[::-1]
            trlist.append((revcomp_seq, i[19]))
    else:
        break

# training data를 excel 파일에 저장하기

from openpyxl import Workbook

write_wb = Workbook()

write_ws = write_wb.active

write_ws['A1'] = 'Promoter'
write_ws['B1'] = 'Reads'

for i in range(len(trlist)):
    write_ws.cell(i+2,1,trlist[i][0])
    write_ws.cell(i+2,2,trlist[i][1])

write_wb.save('/Users/ufooo/OneDrive/문서/Python Scripts/PCC6803 Promoter and reads 50bp_20201215.xlsx')